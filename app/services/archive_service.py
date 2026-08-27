"""
Archive Service — парсинг и обработка выгрузок архива Wildberries (.xlsx).
Обрабатывает листы «КИЗ» и «Сборочные задания» для:
1. Вывода КИЗ из оборота («Дистанционная продажа») с указанием номера и даты кассового чека.
2. Проверки и возврата в оборот отказных КИЗ.
"""
import io
import re
import logging
from datetime import datetime, timezone
from typing import Dict, List, Any, Optional, Tuple

import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.models.order import Order, KizStatus, OrderStatus
from app.models.seller import Seller
from app.models.kiz import KizProductInfo
from app.services.kiz_service import (
    CZ_WITHDRAWAL_STATUSES,
    CZ_STATUS_DESCRIPTIONS,
    is_kiz_withdrawn,
    parse_kiz_code,
    batch_verify_and_sync_cises,
)

logger = logging.getLogger(__name__)


def _parse_date_str(val: Any) -> Optional[str]:
    """Преобразует дату в формат YYYY-MM-DD для True API."""
    if not val:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    
    val_str = str(val).strip()
    # Handle 'HH:MM:SS DD.MM.YYYY' (e.g. '03:04:00 25.08.2026')
    m = re.search(r'(\d{2})\.(\d{2})\.(\d{4})', val_str)
    if m:
        day, month, year = m.groups()
        return f"{year}-{month}-{day}"
    
    # Handle 'YYYY-MM-DD'
    m2 = re.search(r'(\d{4})-(\d{2})-(\d{2})', val_str)
    if m2:
        year, month, day = m2.groups()
        return f"{year}-{month}-{day}"
        
    return None


def parse_wb_archive_excel(file_bytes: bytes) -> Dict[str, List[Dict[str, Any]]]:
    """
    Парсит бинарные данные Excel файла архива WB.
    Возвращает словарь с двумя списками:
    - 'kiz_rows': записи листа «КИЗ»
    - 'tasks_rows': записи листа «Сборочные задания»
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    
    # Locate sheets (case-insensitive search)
    sheet_names = wb.sheetnames
    kiz_sheet_name = None
    tasks_sheet_name = None
    
    for s in sheet_names:
        s_lower = s.strip().lower()
        if "киз" in s_lower:
            kiz_sheet_name = s
        elif "сборочн" in s_lower or "задани" in s_lower or "архив" in s_lower:
            tasks_sheet_name = s
            
    if not kiz_sheet_name and not tasks_sheet_name:
        # Fallback to first sheet
        kiz_sheet_name = sheet_names[0]

    kiz_rows = []
    if kiz_sheet_name and kiz_sheet_name in wb.sheetnames:
        sheet = wb[kiz_sheet_name]
        if sheet.max_row >= 2:
            headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
            headers = [str(h).strip() if h is not None else f"col_{i+1}" for i, h in enumerate(headers)]
            for r in range(2, sheet.max_row + 1):
                row = {}
                is_empty = True
                for c in range(1, sheet.max_column + 1):
                    val = sheet.cell(r, c).value
                    if val is not None:
                        is_empty = False
                    row[headers[c - 1]] = val
                if not is_empty:
                    kiz_rows.append(row)

    tasks_rows = []
    if tasks_sheet_name and tasks_sheet_name in wb.sheetnames:
        sheet = wb[tasks_sheet_name]
        if sheet.max_row >= 2:
            headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
            headers = [str(h).strip() if h is not None else f"col_{i+1}" for i, h in enumerate(headers)]
            for r in range(2, sheet.max_row + 1):
                row = {}
                is_empty = True
                for c in range(1, sheet.max_column + 1):
                    val = sheet.cell(r, c).value
                    if val is not None:
                        is_empty = False
                    row[headers[c - 1]] = val
                if not is_empty:
                    tasks_rows.append(row)

    return {
        "kiz_rows": kiz_rows,
        "tasks_rows": tasks_rows,
    }


async def analyze_archive_data(
    seller: Seller,
    archive_data: Dict[str, List[Dict[str, Any]]],
    db: AsyncSession,
) -> Dict[str, Any]:
    """
    Анализирует распарсенные данные архива WB, сопоставляет с заказами в БД
    и формирует структуру предпросмотра:
    - withdrawals: продажи (с номером чека и датой)
    - returns: возвраты/отказы
    - summary: сводная статистика
    """
    kiz_rows = archive_data.get("kiz_rows", [])
    tasks_rows = archive_data.get("tasks_rows", [])
    
    # Map tasks by order_id or sticker_id
    tasks_by_order_id = {}
    tasks_by_sticker = {}
    for t in tasks_rows:
        order_id = t.get("№ задания") or t.get("Номер задания") or t.get("order_id")
        sticker = t.get("Стикер") or t.get("QR-код поставки") or t.get("sticker_id")
        if order_id:
            try:
                tasks_by_order_id[int(order_id)] = t
            except (ValueError, TypeError):
                tasks_by_order_id[str(order_id)] = t
        if sticker:
            tasks_by_sticker[str(sticker).strip()] = t

    # Extract all order IDs and KIZ codes from kiz_rows
    order_ids = []
    all_kiz_codes = []
    for k in kiz_rows:
        oid = k.get("№ задания") or k.get("Номер задания")
        if oid:
            try:
                order_ids.append(int(oid))
            except (ValueError, TypeError):
                pass
        kiz = str(k.get("КИЗ") or "").strip()
        if kiz:
            all_kiz_codes.append(kiz)

    # 1. Query existing orders in DB
    existing_orders_map: Dict[int, Order] = {}
    if order_ids:
        stmt = select(Order).where(
            Order.seller_id == seller.id,
            Order.id.in_(order_ids)
        )
        res = await db.execute(stmt)
        for ord_obj in res.scalars().all():
            existing_orders_map[ord_obj.id] = ord_obj

    # 2. Query existing KizProductInfo in DB (Single Source of Truth)
    kiz_info_map: Dict[str, KizProductInfo] = {}
    if all_kiz_codes:
        unique_kiz = list(set(all_kiz_codes))
        kiz_stmt = select(KizProductInfo).where(
            (KizProductInfo.seller_id == seller.id) | (KizProductInfo.seller_id.is_(None)),
            KizProductInfo.kiz_code.in_(unique_kiz)
        )
        kiz_res = await db.execute(kiz_stmt)
        for kinfo in kiz_res.scalars().all():
            kiz_info_map[kinfo.kiz_code] = kinfo
            if kinfo.clean_cis:
                kiz_info_map[kinfo.clean_cis] = kinfo

        # 3. Optional batch live verification with True API for missing or unchecked KIZ
        try:
            verified_map = await batch_verify_and_sync_cises(
                seller=seller,
                kiz_codes=unique_kiz,
                db=db,
                force_refresh=False
            )
            for k_code, k_obj in verified_map.items():
                if k_obj:
                    kiz_info_map[k_code] = k_obj
                    if k_obj.clean_cis:
                        kiz_info_map[k_obj.clean_cis] = k_obj
        except Exception as batch_err:
            logger.debug(f"Archive analysis batch verify note: {batch_err}")

    withdrawals = []
    returns = []

    for k in kiz_rows:
        order_id_val = k.get("№ задания") or k.get("Номер задания")
        try:
            order_id = int(order_id_val) if order_id_val is not None else None
        except (ValueError, TypeError):
            order_id = None

        sticker = str(k.get("Стикер") or "").strip()
        kiz_code = str(k.get("КИЗ") or "").strip()
        parsed_kiz = parse_kiz_code(kiz_code)
        clean_cis = parsed_kiz.get("clean_cis") or kiz_code

        receipt_num = str(k.get("Номер чека") or "").strip()
        fn_num = str(k.get("Номер фискального накопителя") or "").strip()
        date_raw = k.get("Дата")
        date_formatted = _parse_date_str(date_raw) or datetime.now(timezone.utc).strftime("%Y-%m-%d")
        op_type = str(k.get("Тип операции") or "").strip().lower()
        price_val = k.get("Стоимость") or 0
        try:
            price = float(price_val)
        except (ValueError, TypeError):
            price = 0.0

        task_info = tasks_by_order_id.get(order_id) or tasks_by_sticker.get(sticker) or {}
        article = task_info.get("Артикул продавца") or task_info.get("Артикул Wildberries") or ""
        name = task_info.get("Наименование") or "Товар WB"
        task_status = task_info.get("Статус задания") or ""

        db_order = existing_orders_map.get(order_id) if order_id else None
        db_kiz_status = db_order.kiz_status.value if db_order and db_order.kiz_status else None
        
        # Look up canonical KizProductInfo
        kinfo_record = kiz_info_map.get(kiz_code) or kiz_info_map.get(clean_cis)
        effective_cz_status = (
            (kinfo_record.cz_status if kinfo_record and kinfo_record.cz_status else None)
            or (db_order.kiz_cz_status if db_order else None)
        )
        effective_status_ex = kinfo_record.cz_status_ex if kinfo_record else None
        raw_payload = kinfo_record.raw_cz_payload if kinfo_record else {}

        # Operation type classification:
        # Приоритет отдается возврату: колонка "Тип операции" == "Возврат", статус задания "отказ"/"отмена"
        # или статус заказа в БД CANCELLED. У возвратов тоже может быть номер чека (чек возврата),
        # поэтому наличие чека не должно автоматически классифицировать строку как продажу.
        is_return = (
            "возврат" in op_type
            or "отказ" in task_status.lower()
            or "отмен" in task_status.lower()
            or (db_order and db_order.status == OrderStatus.CANCELLED)
        )
        is_sale = not is_return and (
            "продаж" in op_type
            or "продано" in task_status.lower()
            or "sold" in task_status.lower()
            or bool(receipt_num)
        )

        # Check withdrawal status from True API specification
        withdrawn_flag, withdraw_reason = is_kiz_withdrawn(
            status=effective_cz_status,
            status_ex=effective_status_ex,
            raw_payload=raw_payload
        )
        is_already_withdrawn = withdrawn_flag or (db_kiz_status == KizStatus.WITHDRAWN.value)

        cz_status_desc = CZ_STATUS_DESCRIPTIONS.get(
            effective_cz_status or "", 
            effective_cz_status or ("Выбыл" if is_already_withdrawn else "Не проверен")
        )

        if is_sale:
            # Check if needs withdrawal: not yet WITHDRAWN / RETIRED
            needs_withdrawal = not is_already_withdrawn

            withdrawals.append({
                "order_id": order_id,
                "sticker_id": sticker,
                "kiz_code": kiz_code,
                "receipt_number": receipt_num,
                "fn_number": fn_num,
                "receipt_date": date_formatted,
                "price": price,
                "price_kopecks": int(round(price * 100)),
                "article": article,
                "name": name,
                "task_status": task_status or "Продано",
                "db_status": db_order.status.value if db_order else "Не найден в БД",
                "db_kiz_status": db_kiz_status or "Не привязан",
                "cz_status": effective_cz_status,
                "cz_status_desc": cz_status_desc,
                "is_already_withdrawn": is_already_withdrawn,
                "needs_withdrawal": needs_withdrawal,
                "selected": needs_withdrawal and bool(kiz_code),
            })
        elif is_return:
            # For returns: needs CZ return only if it was previously withdrawn and not yet returned
            needs_cz_return = is_already_withdrawn

            returns.append({
                "order_id": order_id,
                "sticker_id": sticker,
                "kiz_code": kiz_code,
                "receipt_number": receipt_num,
                "fn_number": fn_num,
                "receipt_date": date_formatted,
                "price": price,
                "price_kopecks": int(round(price * 100)),
                "article": article,
                "name": name,
                "task_status": task_status or "Возврат / отказ покупателя",
                "db_status": db_order.status.value if db_order else "Не найден в БД",
                "db_kiz_status": db_kiz_status or "Не привязан",
                "cz_status": effective_cz_status,
                "db_cz_status": effective_cz_status,
                "cz_status_desc": cz_status_desc,
                "needs_cz_return": needs_cz_return,
                "action_recommended": "⚠️ Требует возврата в оборот" if needs_cz_return else "✅ Уже в обороте (готов к привязке)",
                "selected": needs_cz_return and bool(kiz_code),
            })

    return {
        "withdrawals": withdrawals,
        "returns": returns,
        "summary": {
            "total_rows": len(kiz_rows),
            "sales_count": len(withdrawals),
            "sales_needing_withdrawal": sum(1 for w in withdrawals if w["needs_withdrawal"]),
            "returns_count": len(returns),
            "returns_needing_cz_return": sum(1 for r in returns if r["needs_cz_return"]),
        }
    }
